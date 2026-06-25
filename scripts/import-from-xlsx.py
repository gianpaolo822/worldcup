#!/usr/bin/env python3
"""从 data/import/ 下的 Excel 导入球队、赛程与历届奖项 JSON。"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
IMPORT_DIR = ROOT / "data" / "import"
DATA_DIR = ROOT / "data"
GENERATED_DIR = DATA_DIR / "generated"

WORLD_CUP_XLSX = IMPORT_DIR / "2026FIFA World Cup.xlsx"
AWARDS_XLSX = IMPORT_DIR / "世界杯奖项体系汇总.xlsx"
OFFICIAL_KICKOFFS_JSON = DATA_DIR / "official-kickoff-times.json"

BEIJING = timezone(timedelta(hours=8))
YEAR = 2026

FIFA_MEMBERS_JSON = DATA_DIR / "fifa-members.json"

FLAGS: dict[str, str] = {
    "墨西哥": "🇲🇽", "南非": "🇿🇦", "韩国": "🇰🇷", "捷克": "🇨🇿", "捷克共和国": "🇨🇿",
    "加拿大": "🇨🇦", "波黑": "🇧🇦", "波斯尼亚和黑塞哥维那": "🇧🇦", "卡塔尔": "🇶🇦", "瑞士": "🇨🇭",
    "巴西": "🇧🇷", "摩洛哥": "🇲🇦", "海地": "🇭🇹", "苏格兰": "🏴󠁧󠁢󠁳󠁣󠁴󠁿",
    "美国": "🇺🇸", "巴拉圭": "🇵🇾", "澳大利亚": "🇦🇺", "土耳其": "🇹🇷",
    "德国": "🇩🇪", "库拉索": "🇨🇼", "科特迪瓦": "🇨🇮", "厄瓜多尔": "🇪🇨",
    "荷兰": "🇳🇱", "日本": "🇯🇵", "瑞典": "🇸🇪", "突尼斯": "🇹🇳",
    "比利时": "🇧🇪", "埃及": "🇪🇬", "伊朗": "🇮🇷", "新西兰": "🇳🇿",
    "西班牙": "🇪🇸", "佛得角": "🇨🇻", "沙特阿拉伯": "🇸🇦", "乌拉圭": "🇺🇾",
    "法国": "🇫🇷", "塞内加尔": "🇸🇳", "伊拉克": "🇮🇶", "挪威": "🇳🇴",
    "阿根廷": "🇦🇷", "阿尔及利亚": "🇩🇿", "奥地利": "🇦🇹", "约旦": "🇯🇴",
    "葡萄牙": "🇵🇹", "民主刚果": "🇨🇩", "刚果民主共和国": "🇨🇩", "乌兹别克斯坦": "🇺🇿", "哥伦比亚": "🇨🇴",
    "英格兰": "🏴󠁧󠁢󠁥󠁮󠁧󠁿", "克罗地亚": "🇭🇷", "加纳": "🇬🇭", "巴拿马": "🇵🇦",
}

CODES: dict[str, str] = {
    "Mexico": "MEX", "South Africa": "RSA", "South Korea": "KOR", "Korea Republic": "KOR",
    "Czechia": "CZE", "Czech Republic": "CZE",
    "Canada": "CAN", "Bosnia and Herzegovina": "BIH", "Qatar": "QAT", "Switzerland": "SUI",
    "Brazil": "BRA", "Morocco": "MAR", "Haiti": "HAI", "Scotland": "SCO",
    "United States": "USA", "Paraguay": "PAR", "Australia": "AUS", "Turkey": "TUR",
    "Germany": "GER", "Curacao": "CUW", "Curaçao": "CUW", "Ivory Coast": "CIV",
    "Côte d'Ivoire": "CIV", "Ecuador": "ECU",
    "Netherlands": "NED", "Japan": "JPN", "Sweden": "SWE", "Tunisia": "TUN",
    "Belgium": "BEL", "Egypt": "EGY", "Iran": "IRN", "New Zealand": "NZL",
    "Spain": "ESP", "Cape Verde": "CPV", "Saudi Arabia": "KSA", "Uruguay": "URU",
    "France": "FRA", "Senegal": "SEN", "Iraq": "IRQ", "Norway": "NOR",
    "Argentina": "ARG", "Algeria": "ALG", "Austria": "AUT", "Jordan": "JOR",
    "Portugal": "POR", "DR Congo": "COD", "Congo DR": "COD", "Uzbekistan": "UZB", "Colombia": "COL",
    "England": "ENG", "Croatia": "CRO", "Ghana": "GHA", "Panama": "PAN",
}

# Excel 中文全称 → 项目内简称（见 scripts/apply-fifa-names.ts NAME_ZH_SHORT）
ZH_ALIASES: dict[str, str] = {
    "捷克共和国": "捷克",
    "波斯尼亚和黑塞哥维那": "波黑",
    "刚果民主共和国": "民主刚果",
}

NAME_ZH_SHORT: dict[str, str] = {
    "CZE": "捷克",
    "BIH": "波黑",
    "COD": "民主刚果",
}

COLORS = [
    "from-sky-400 to-sky-600",
    "from-blue-700 to-blue-900",
    "from-red-600 to-blue-700",
    "from-green-500 to-red-500",
]

CITY_SHORT: dict[str, str] = {
    "墨西哥城": "MEX", "萨波潘": "GDL", "多伦多": "TOR", "温哥华": "VAN",
    "洛杉矶": "LA", "圣克拉拉": "SJ", "纽约/新泽西": "NY", "福克斯堡": "BOS",
    "达拉斯": "DAL", "亚特兰大": "ATL", "休斯顿": "HOU", "休斯敦": "HOU",
    "堪萨斯城": "KC", "费城": "PHI", "西雅图": "SEA", "迈阿密": "MIA", "阿灵顿": "ARL",
    "蒙特雷": "MTY", "波士顿": "BOS",
}

# 球场城市 → IANA 时区（从当地时间换算 UTC / 北京时间，勿用 Excel「北京时间」列）
CITY_TIMEZONE: dict[str, str] = {
    "墨西哥城": "America/Mexico_City",
    "萨波潘": "America/Mexico_City",
    "蒙特雷": "America/Monterrey",
    "多伦多": "America/Toronto",
    "温哥华": "America/Vancouver",
    "洛杉矶": "America/Los_Angeles",
    "圣克拉拉": "America/Los_Angeles",
    "西雅图": "America/Los_Angeles",
    "纽约/新泽西": "America/New_York",
    "费城": "America/New_York",
    "福克斯堡": "America/New_York",
    "波士顿": "America/New_York",
    "迈阿密": "America/New_York",
    "亚特兰大": "America/New_York",
    "达拉斯": "America/Chicago",
    "阿灵顿": "America/Chicago",
    "休斯顿": "America/Chicago",
    "休斯敦": "America/Chicago",
    "堪萨斯城": "America/Chicago",
}

# Excel「球员名单 / 因伤退出」里的英文队名与「分组」表 nameEn 不一致时的映射
TEAM_EN_ALIASES: dict[str, str] = {
    "USA": "United States",
    "Curacao": "Curaçao",
    "Côte D'Ivoire": "Côte d'Ivoire",
    "Cote D'Ivoire": "Côte d'Ivoire",
    "Côte d'Ivoire": "Côte d'Ivoire",
    "Ivory Coast": "Côte d'Ivoire",
    "Czechia": "Czech Republic",
    "DR Congo": "Congo DR",
    "South Korea": "Korea Republic",
}


def load_fifa_by_code() -> dict[str, dict]:
    if not FIFA_MEMBERS_JSON.exists():
        return {}
    data = json.loads(FIFA_MEMBERS_JSON.read_text(encoding="utf-8"))
    return {m["code"]: m for m in data.get("members", [])}


def apply_fifa_names(teams: list[dict]) -> list[dict]:
    fifa = load_fifa_by_code()
    if not fifa:
        return teams
    updated: list[dict] = []
    for team in teams:
        member = fifa.get(team["code"])
        if member:
            team = {
                **team,
                "name": NAME_ZH_SHORT.get(member["code"], member["nameZh"]),
                "nameEn": member["nameEn"],
                "code": member["code"],
            }
        updated.append(team)
    return updated


def build_teams_by_zh(teams: list[dict]) -> dict[str, dict]:
    by_zh = {t["name"]: t for t in teams}
    for short, canonical in ZH_ALIASES.items():
        if canonical in by_zh:
            by_zh[short] = by_zh[canonical]
    return by_zh


def resolve_team_by_en(country_en: str, teams_by_en: dict[str, dict]) -> dict | None:
    text = country_en.strip()
    if not text:
        return None
    if text in teams_by_en:
        return teams_by_en[text]
    canonical = TEAM_EN_ALIASES.get(text)
    if canonical and canonical in teams_by_en:
        return teams_by_en[canonical]
    folded = text.casefold()
    for name, team in teams_by_en.items():
        if name.casefold() == folded:
            return team
    return None


def slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return s or "team"


def zh_only(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if "\n" in text:
        text = text.split("\n", 1)[0].strip()
    return text


def parse_cn_date(text: object) -> tuple[int, int] | None:
    if not text:
        return None
    m = re.search(r"(\d+)\s*月\s*(\d+)\s*日", str(text))
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def parse_local_time(time_text: object) -> tuple[int, int]:
    if isinstance(time_text, datetime):
        return time_text.hour, time_text.minute
    if isinstance(time_text, (int, float)):
        total_minutes = int(round(float(time_text) * 24 * 60)) % (24 * 60)
        return total_minutes // 60, total_minutes % 60
    time_str = str(time_text or "00:00").strip()
    hour, minute = (int(x) for x in time_str.split(":")[:2])
    return hour, minute


def normalize_local_time(time_text: object) -> str:
    hour, minute = parse_local_time(time_text)
    return f"{hour:02d}:{minute:02d}"


def load_official_kickoffs() -> tuple[dict[int, dict], dict[tuple[str, str, str], dict]]:
    if not OFFICIAL_KICKOFFS_JSON.exists():
        return {}, {}
    data = json.loads(OFFICIAL_KICKOFFS_JSON.read_text(encoding="utf-8"))
    by_seq: dict[int, dict] = {}
    by_slot: dict[tuple[str, str, str], dict] = {}
    for m in data.get("matches", []):
        seq = int(m["seq"])
        by_seq[seq] = m
        slot = (str(m["cityZh"]), str(m["localDate"]), str(m["localTime"]))
        by_slot[slot] = m
    return by_seq, by_slot


def resolve_official_kickoff(
    seq: int,
    city: object,
    local_date: object,
    local_time: object,
    official_by_seq: dict[int, dict],
    official_by_slot: dict[tuple[str, str, str], dict],
) -> dict | None:
    """优先按城市+当地日期+时间匹配 FIFA 官方表；序号仅在同城时作兜底。"""
    city_str = str(city or "").strip()
    date_str = str(local_date or "").strip()
    time_str = normalize_local_time(local_time)
    slot_hit = official_by_slot.get((city_str, date_str, time_str))
    if slot_hit:
        return slot_hit
    seq_hit = official_by_seq.get(seq)
    if seq_hit and seq_hit.get("cityZh") == city_str:
        return seq_hit
    return None


def parse_local_datetime(city: object, date_text: object, time_text: object) -> tuple[str, str]:
    """从球场当地日期/时间换算为北京时间展示串与 UTC kickoffAt。"""
    parsed = parse_cn_date(date_text)
    if not parsed:
        return ("日期待定", "")
    month, day = parsed
    hour, minute = parse_local_time(time_text)
    city_str = str(city or "").strip()
    tz_name = CITY_TIMEZONE.get(city_str, "America/Mexico_City")
    local_dt = datetime(YEAR, month, day, hour, minute, tzinfo=ZoneInfo(tz_name))
    utc_dt = local_dt.astimezone(timezone.utc)
    beijing_dt = local_dt.astimezone(BEIJING)
    display = beijing_dt.strftime("%Y-%m-%d %H:%M")
    kickoff = utc_dt.isoformat().replace("+00:00", "Z")
    return display, kickoff


def official_kickoff_display(official: dict) -> tuple[str, str]:
    return str(official["beijingTime"]), str(official["kickoffAt"])


def parse_beijing_datetime(date_text: object, time_text: object) -> tuple[str, str]:
    """兼容旧逻辑；赛程导入已改用 parse_local_datetime。"""
    parsed = parse_cn_date(date_text)
    time_str = str(time_text or "00:00").strip()
    if not parsed:
        return ("日期待定", "")
    month, day = parsed
    display = f"{YEAR}-{month:02d}-{day:02d} {time_str}"
    hour, minute = parse_local_time(time_text)
    dt = datetime(YEAR, month, day, hour, minute, tzinfo=BEIJING)
    return display, dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def load_teams(wb) -> list[dict]:
    ws = wb["分组"]
    teams: list[dict] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[1]:
            continue
        zh, en, group = str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()
        teams.append(
            {
                "id": slugify(en),
                "name": zh,
                "nameEn": en,
                "code": CODES.get(en, en[:3].upper()),
                "group": group,
                "flag": FLAGS.get(zh, "🏳️"),
                "color": COLORS[i % len(COLORS)],
            }
        )
    return teams


def resolve_team(name: object, group: str, teams_by_zh: dict[str, dict]) -> dict:
    label = str(name or "").strip()
    label = ZH_ALIASES.get(label, label)
    if label in teams_by_zh:
        return teams_by_zh[label]
    return {
        "id": slugify(label),
        "name": label,
        "nameEn": label,
        "code": label[:3].upper(),
        "group": group or "",
        "flag": "🏳️",
        "color": "from-neutral-600 to-neutral-800",
    }


def load_matches(
    wb,
    teams_by_zh: dict[str, dict],
    official_by_seq: dict[int, dict],
    official_by_slot: dict[tuple[str, str, str], dict],
) -> list[dict]:
    ws = wb["赛程"]
    matches: list[dict] = []
    group_matchday: dict[str, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        seq = int(row[0])
        home_name, away_name = row[3], row[5]
        city, venue = row[6], row[7]
        stage_raw = str(row[8] or "小组赛").strip()
        group = str(row[9] or "").strip()
        local_date, local_time = row[1], row[2]

        official = resolve_official_kickoff(
            seq, city, local_date, local_time, official_by_seq, official_by_slot
        )
        if official:
            time_display, kickoff_at = official_kickoff_display(official)
        else:
            time_display, kickoff_at = parse_local_datetime(city, local_date, local_time)
        home = resolve_team(home_name, group, teams_by_zh)
        away = resolve_team(away_name, group, teams_by_zh)

        matchday = 1
        if group and stage_raw == "小组赛":
            group_matchday[group] = group_matchday.get(group, 0) + 1
            matchday = group_matchday[group]

        stage = stage_raw
        if stage_raw == "1/4决赛":
            stage = "四分之一决赛"
        elif stage_raw == "半决赛":
            stage = "半决赛"
        elif stage_raw == "三四名决赛":
            stage = "季军战"
        elif stage_raw == "决赛":
            stage = "决赛"

        city_str = str(city or "").strip()
        venue_str = str(venue or "").strip()
        match = {
            "id": f"wc2026-m{seq:03d}",
            "homeTeam": home,
            "awayTeam": away,
            "homeScore": None,
            "awayScore": None,
            "status": "upcoming",
            "time": time_display,
            "kickoffAt": kickoff_at or None,
            "venue": venue_str,
            "venueShort": CITY_SHORT.get(city_str, city_str[:3].upper() if city_str else "—"),
            "venueCity": city_str or None,
            "matchday": matchday,
            "group": group or None,
            "stage": stage,
        }
        matches.append({k: v for k, v in match.items() if v is not None or k in ("homeScore", "awayScore", "kickoffAt")})
    return matches


def excel_serial_to_date(value: object) -> str | None:
    if value is None:
        return None
    try:
        serial = float(value)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=serial)).strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def parse_player_dob(value: object) -> str | None:
    if not value:
        return None
    text = str(value).strip()
    m = re.match(r"(\d{2})-(\d{2})-(\d{2})", text)
    if not m:
        return text
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    full_year = 1900 + year if year >= 30 else 2000 + year
    return f"{full_year:04d}-{month:02d}-{day:02d}"


def load_injured_keys(wb, teams_by_en: dict[str, dict]) -> dict[str, str]:
    if "因伤退出" not in wb.sheetnames:
        return {}
    ws = wb["因伤退出"]
    injured: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        country_en = str(row[9] or "").strip()
        number = int(row[0])
        team = resolve_team_by_en(country_en, teams_by_en)
        team_id = team["id"] if team else slugify(country_en)
        key = f"{team_id}-{number}"
        note = str(row[13] or "").strip() if len(row) > 13 else ""
        injured[key] = note or "因伤退出世界杯"
    return injured


def load_players(wb, teams_by_en: dict[str, dict], injured: dict[str, str]) -> list[dict]:
    ws = wb["球员名单"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell or "").strip() for cell in header_row]
    header_index = {label: idx for idx, label in enumerate(headers) if label}

    def col(label: str, fallback: int) -> int:
        return header_index[label] if label in header_index else fallback

    name_zh_idx = col("中文名", -1)
    club_zh_idx = col("中文俱乐部", -1)

    players: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        number = int(row[0])
        name_en = str(row[1]).strip()
        nat = str(row[2] or "").strip()
        position = str(row[3] or "").strip()
        height = row[4]
        weight = row[5]
        dob = parse_player_dob(row[6])
        birth_place = str(row[7] or "").strip()
        club_en = str(row[8] or "").strip()
        country_en = str(row[9] or "").strip()
        group = str(row[10] or "").strip()

        name_zh = ""
        if name_zh_idx >= 0 and len(row) > name_zh_idx and row[name_zh_idx]:
            name_zh = str(row[name_zh_idx]).strip()
        club_zh = ""
        if club_zh_idx >= 0 and len(row) > club_zh_idx and row[club_zh_idx]:
            club_zh = str(row[club_zh_idx]).strip()

        team = resolve_team_by_en(country_en, teams_by_en)
        team_id = team["id"] if team else slugify(country_en)
        player_id = f"{team_id}-{number}"

        player = {
            "id": player_id,
            "teamId": team_id,
            "number": number,
            "nameEn": name_en,
            "nat": nat,
            "position": position,
            "height": float(height) if height else None,
            "weight": float(weight) if weight else None,
            "birthDate": dob,
            "birthPlace": birth_place or None,
            "clubEn": club_en or None,
            "group": group or None,
            "status": "injured" if player_id in injured else "active",
        }
        if name_zh:
            player["nameZh"] = name_zh
        if club_zh:
            player["club"] = club_zh
        if player_id in injured:
            player["injuryNote"] = injured[player_id]
        players.append({k: v for k, v in player.items() if v is not None})
    return players


def load_coaches(wb, teams_by_zh: dict[str, dict]) -> list[dict]:
    ws = wb["主教练"]
    coaches: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        team_zh = str(row[1]).strip()
        team_en = str(row[2] or "").strip()
        name_zh = str(row[3] or "").strip()
        name_en = str(row[4] or "").strip()
        birth = excel_serial_to_date(row[5])
        nationality = str(row[6] or "").strip()
        group = str(row[7] or "").strip()

        team = teams_by_zh.get(team_zh)
        team_id = team["id"] if team else slugify(team_en or team_zh)

        coach = {
            "teamId": team_id,
            "name": name_zh or name_en,
            "nameEn": name_en or name_zh,
            "nationality": nationality or None,
            "birthDate": birth,
            "group": group or None,
        }
        coaches.append({k: v for k, v in coach.items() if v is not None})
    return coaches


def load_venues(wb) -> list[dict]:
    ws = wb["体育场"]
    venues: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        country_zh = str(row[1] or "").strip()
        country_en = str(row[2] or "").strip()
        city_zh = str(row[3] or "").strip()
        city_en = str(row[4] or "").strip()
        name_zh = str(row[5] or "").strip()
        name_en = str(row[6] or "").strip()
        capacity = row[7]
        match_count = row[8]
        note = str(row[9] or "").strip() if len(row) > 9 else ""

        venue = {
            "id": slugify(name_zh or name_en),
            "name": name_zh or name_en,
            "nameEn": name_en or name_zh,
            "city": city_zh or city_en,
            "cityEn": city_en or city_zh,
            "country": country_zh or country_en,
            "countryEn": country_en or country_zh,
            "capacity": int(capacity) if capacity else None,
            "matchCount": int(match_count) if match_count else None,
            "note": note or None,
        }
        venues.append({k: v for k, v in venue.items() if v is not None})
    return venues


def load_world_cup_history(wb) -> list[dict]:
    champions = wb["冠亚军 Champions"]
    golden_ball = {
        int(r[0]): zh_only(r[3])
        for r in wb["金球奖 Golden Ball"].iter_rows(min_row=2, values_only=True)
        if r and r[0]
    }
    golden_boot: dict[int, str] = {}
    for r in wb["金靴奖 Golden Boot"].iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        year = int(r[0])
        winner = zh_only(r[2])
        goals = r[3]
        if winner and goals not in (None, "—"):
            golden_boot[year] = f"{winner} ({goals}球)"
        elif winner:
            golden_boot[year] = winner
    golden_glove = {
        int(r[0]): zh_only(r[2])
        for r in wb["金手套奖 Golden Glove"].iter_rows(min_row=2, values_only=True)
        if r and r[0]
    }

    history: list[dict] = []
    for row in champions.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        year = int(row[0])
        history.append(
            {
                "year": year,
                "host": zh_only(row[6]),
                "champion": zh_only(row[2]),
                "runnerUp": zh_only(row[3]),
                "score": "",
                "goldenBoot": golden_boot.get(year, "—"),
                "goldenBall": golden_ball.get(year, "—"),
                "goldenGlove": golden_glove.get(year, "—"),
            }
        )
    history.sort(key=lambda x: x["year"], reverse=True)
    return history


def write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    if not WORLD_CUP_XLSX.exists():
        raise SystemExit(f"缺少文件: {WORLD_CUP_XLSX}")
    if not AWARDS_XLSX.exists():
        raise SystemExit(f"缺少文件: {AWARDS_XLSX}")

    wb = load_workbook(WORLD_CUP_XLSX, read_only=True, data_only=True)
    teams = apply_fifa_names(load_teams(wb))
    teams_by_zh = build_teams_by_zh(teams)
    teams_by_en = {t["nameEn"]: t for t in teams}
    injured = load_injured_keys(wb, teams_by_en)
    players = load_players(wb, teams_by_en, injured)
    coaches = load_coaches(wb, teams_by_zh)
    venues = load_venues(wb)
    official_by_seq, official_by_slot = load_official_kickoffs()
    matches = load_matches(wb, teams_by_zh, official_by_seq, official_by_slot)
    wb.close()

    awards_wb = load_workbook(AWARDS_XLSX, read_only=True, data_only=True)
    history = load_world_cup_history(awards_wb)
    awards_wb.close()

    meta = {
        "syncedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source": "data/import/*.xlsx",
        "note": "由 scripts/import-from-xlsx.py 从用户整理 Excel 导入",
    }

    write_json(GENERATED_DIR / "teams.json", teams)
    write_json(GENERATED_DIR / "matches.json", matches)
    write_json(GENERATED_DIR / "players.json", players)
    write_json(GENERATED_DIR / "sync-meta.json", meta)
    write_json(DATA_DIR / "teams.json", teams)
    write_json(DATA_DIR / "coaches.json", coaches)
    write_json(DATA_DIR / "venues.json", venues)
    write_json(DATA_DIR / "world-cup-history.json", history)

    print(
        f"[data:import] teams={len(teams)} matches={len(matches)} "
        f"players={len(players)} coaches={len(coaches)} venues={len(venues)} history={len(history)}"
    )


if __name__ == "__main__":
    main()
